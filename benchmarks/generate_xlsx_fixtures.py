#!/usr/bin/env python3
"""
Generate XLSX test fixtures for import performance benchmarking.

Creates files of varying sizes:
- small.xlsx: ~5k cells
- medium.xlsx: ~200k cells, some formulas
- large.xlsx: ~1M cells, mixed content

Files are placed in benchmarks/fixtures/ (gitignored).
Run: python3 benchmarks/generate_xlsx_fixtures.py
"""

import os
import random
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    exit(1)

FIXTURES_DIR = Path(__file__).parent / "fixtures"


def generate_small_xlsx():
    """Generate small.xlsx with ~5,000 cells."""
    print("Generating small.xlsx (~5k cells)...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # 100 rows x 50 cols = 5,000 cells
    for row in range(1, 101):
        for col in range(1, 51):
            if col == 1:
                ws.cell(row=row, column=col, value=f"Row {row}")
            elif col == 2:
                ws.cell(row=row, column=col, value=row * 10)
            else:
                ws.cell(row=row, column=col, value=random.random() * 1000)

    filepath = FIXTURES_DIR / "small.xlsx"
    wb.save(filepath)
    print(f"  Created: {filepath} ({filepath.stat().st_size:,} bytes)")


def generate_medium_xlsx():
    """Generate medium.xlsx with ~200,000 cells and formulas."""
    print("Generating medium.xlsx (~200k cells with formulas)...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Header row
    headers = ["ID", "Region", "Product", "Quantity", "Price", "Total", "Tax", "Grand Total"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    regions = ["North", "South", "East", "West", "Central"]
    products = ["Widget", "Gadget", "Gizmo", "Doohickey", "Thingamajig"]

    # Data rows: 25,000 rows x 8 cols = 200,000 cells
    for row in range(2, 25002):
        ws.cell(row=row, column=1, value=row - 1)  # ID
        ws.cell(row=row, column=2, value=random.choice(regions))  # Region
        ws.cell(row=row, column=3, value=random.choice(products))  # Product
        ws.cell(row=row, column=4, value=random.randint(1, 100))  # Quantity
        ws.cell(row=row, column=5, value=round(random.uniform(10, 500), 2))  # Price
        # Formulas
        ws.cell(row=row, column=6, value=f"=D{row}*E{row}")  # Total
        ws.cell(row=row, column=7, value=f"=F{row}*0.08")  # Tax
        ws.cell(row=row, column=8, value=f"=F{row}+G{row}")  # Grand Total

    # Add a second sheet with summary formulas
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Total Records"
    ws2["B1"] = "=COUNTA(Sales!A:A)-1"
    ws2["A2"] = "Sum of Totals"
    ws2["B2"] = "=SUM(Sales!F:F)"
    ws2["A3"] = "Average Price"
    ws2["B3"] = "=AVERAGE(Sales!E:E)"

    filepath = FIXTURES_DIR / "medium.xlsx"
    wb.save(filepath)
    print(f"  Created: {filepath} ({filepath.stat().st_size:,} bytes)")


def generate_large_xlsx():
    """Generate large.xlsx with ~1,000,000 cells."""
    print("Generating large.xlsx (~1M cells)...")
    wb = Workbook()

    # Multiple sheets to spread the load
    # 4 sheets x 250,000 cells each = 1,000,000 cells
    # Each sheet: 5000 rows x 50 cols

    for sheet_num in range(4):
        if sheet_num == 0:
            ws = wb.active
            ws.title = f"Data_{sheet_num + 1}"
        else:
            ws = wb.create_sheet(f"Data_{sheet_num + 1}")

        print(f"  Generating sheet {sheet_num + 1}/4...")

        # Header
        for col in range(1, 51):
            ws.cell(row=1, column=col, value=f"Col_{col}")

        # Data: 5000 rows x 50 cols = 250,000 cells per sheet
        for row in range(2, 5002):
            for col in range(1, 51):
                # Mix of types
                if col == 1:
                    ws.cell(row=row, column=col, value=f"S{sheet_num+1}R{row}")
                elif col % 5 == 0:
                    # Some formulas
                    prev_col = get_column_letter(col - 1)
                    ws.cell(row=row, column=col, value=f"={prev_col}{row}*2")
                elif col % 3 == 0:
                    ws.cell(row=row, column=col, value=f"Text_{row}_{col}")
                else:
                    ws.cell(row=row, column=col, value=random.random() * 10000)

    filepath = FIXTURES_DIR / "large.xlsx"
    wb.save(filepath)
    print(f"  Created: {filepath} ({filepath.stat().st_size:,} bytes)")


def main():
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Output directory: {FIXTURES_DIR}")
    print()

    generate_small_xlsx()
    generate_medium_xlsx()
    generate_large_xlsx()

    print()
    print("Done! Run benchmarks with:")
    print("  cargo run -p visigrid-gpui --release -- benchmarks/fixtures/small.xlsx")


if __name__ == "__main__":
    main()
